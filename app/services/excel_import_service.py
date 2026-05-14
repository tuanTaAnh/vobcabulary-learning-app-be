import io
from typing import Optional

from openpyxl import load_workbook
from sqlmodel import Session, select

from app.models.models import Collection, Vocab


HEADER_ROW = 1

COL_GERMAN = "german"
COL_PRONUNCIATION = "pronunciation"
COL_MEANING = "meaning"
COL_EXAMPLES = "examples"


def normalize_text(value) -> str:
    if value is None:
        return ""

    return str(value).strip()


def get_first_line(value) -> str:
    text = normalize_text(value)
    lines = [line.strip() for line in text.splitlines() if line.strip()]

    if not lines:
        return ""

    return lines[0]


def clean_examples(value) -> str:
    text = normalize_text(value)

    if not text:
        return ""

    lines = text.splitlines()
    cleaned_lines = []

    for line in lines:
        stripped = line.strip()

        if not stripped:
            continue

        if stripped.lower().startswith("note"):
            break

        cleaned_lines.append(stripped)

    return "\n".join(cleaned_lines).strip()


def combine_meaning_and_pronunciation(meaning: str, pronunciation: str) -> str:
    meaning = normalize_text(meaning)
    pronunciation = normalize_text(pronunciation)

    if meaning and pronunciation:
        return f"{meaning} /{pronunciation}/"

    if meaning:
        return meaning

    if pronunciation:
        return f"/{pronunciation}/"

    return ""


def find_header_map(sheet) -> dict[str, int]:
    header_map = {}

    for cell in sheet[HEADER_ROW]:
        header = normalize_text(cell.value).lower()

        if header:
            header_map[header] = cell.column

    required_headers = [
        COL_GERMAN,
        COL_PRONUNCIATION,
        COL_MEANING,
        COL_EXAMPLES,
    ]

    missing_headers = [
        header for header in required_headers if header not in header_map
    ]

    if missing_headers:
        raise ValueError(
            "Missing required Excel headers: "
            + ", ".join(missing_headers)
            + ". Expected headers: German, Pronunciation, meaning, Examples"
        )

    return header_map


def get_or_create_collection(
    session: Session,
    collection_name: str,
    description: Optional[str] = None,
    icon: Optional[str] = None,
) -> Collection:
    statement = select(Collection).where(Collection.name == collection_name)
    collection = session.exec(statement).first()

    if collection:
        changed = False

        if description is not None and collection.description != description:
            collection.description = description
            changed = True

        if hasattr(collection, "icon") and icon and collection.icon != icon:
            collection.icon = icon
            changed = True

        if changed:
            session.add(collection)
            session.commit()
            session.refresh(collection)

        return collection

    create_data = {
        "name": collection_name,
        "description": description or "Imported from Excel",
    }

    if hasattr(Collection, "icon"):
        create_data["icon"] = icon or "🇩🇪"

    collection = Collection(**create_data)

    session.add(collection)
    session.commit()
    session.refresh(collection)

    return collection


def get_existing_vocab(
    session: Session,
    german: str,
    collection_id: int,
) -> Optional[Vocab]:
    statement = (
        select(Vocab)
        .where(Vocab.german == german)
        .where(Vocab.collection_id == collection_id)
    )

    return session.exec(statement).first()


def import_excel_to_collection(
    session: Session,
    file_bytes: bytes,
    collection_name: str = "German Vocab",
    collection_description: str = "Imported from Excel",
    collection_icon: str = "🇩🇪",
    default_topic: Optional[str] = None,
    mode: str = "update",
    sheet_name: Optional[str] = None,
) -> dict:
    if mode not in {"update", "skip"}:
        raise ValueError("mode must be either 'update' or 'skip'.")

    workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)

    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' not found. Available sheets: {workbook.sheetnames}"
            )

        sheet = workbook[sheet_name]
    else:
        sheet = workbook.active

    header_map = find_header_map(sheet)

    collection = get_or_create_collection(
        session=session,
        collection_name=collection_name,
        description=collection_description,
        icon=collection_icon,
    )

    if collection.id is None:
        raise RuntimeError("Collection ID is missing after creation.")

    imported_count = 0
    updated_count = 0
    skipped_count = 0

    for row_index in range(HEADER_ROW + 1, sheet.max_row + 1):
        german_raw = sheet.cell(
            row=row_index,
            column=header_map[COL_GERMAN],
        ).value

        pronunciation_raw = sheet.cell(
            row=row_index,
            column=header_map[COL_PRONUNCIATION],
        ).value

        meaning_raw = sheet.cell(
            row=row_index,
            column=header_map[COL_MEANING],
        ).value

        examples_raw = sheet.cell(
            row=row_index,
            column=header_map[COL_EXAMPLES],
        ).value

        german = get_first_line(german_raw)
        pronunciation = normalize_text(pronunciation_raw)
        meaning = normalize_text(meaning_raw)
        examples = clean_examples(examples_raw)

        if not german:
            skipped_count += 1
            continue

        vietnamese = combine_meaning_and_pronunciation(
            meaning=meaning,
            pronunciation=pronunciation,
        )

        existing_vocab = get_existing_vocab(
            session=session,
            german=german,
            collection_id=collection.id,
        )

        if existing_vocab:
            if mode == "skip":
                skipped_count += 1
                continue

            existing_vocab.vietnamese = vietnamese
            existing_vocab.examples = examples
            existing_vocab.topic = default_topic
            existing_vocab.collection_id = collection.id

            session.add(existing_vocab)
            updated_count += 1
        else:
            vocab = Vocab(
                german=german,
                vietnamese=vietnamese,
                examples=examples,
                topic=default_topic,
                collection_id=collection.id,
                swipe_count=0,
                is_starred=False,
            )

            session.add(vocab)
            imported_count += 1

    session.commit()

    return {
        "success": True,
        "sheet": sheet.title,
        "collection_id": collection.id,
        "collection_name": collection.name,
        "imported": imported_count,
        "updated": updated_count,
        "skipped": skipped_count,
        "total_processed": imported_count + updated_count + skipped_count,
    }
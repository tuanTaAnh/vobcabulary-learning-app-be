import os
import sys
from pathlib import Path
from typing import Optional

# =========================
# Hardcoded settings
# =========================

BACKEND_ROOT = Path(__file__).resolve().parents[2]

EXCEL_PATH = BACKEND_ROOT / "data" / "german_vocab.xlsx"
DATABASE_PATH = BACKEND_ROOT / "data" / "vocab.db"

COLLECTION_NAME = "German Vocab"
COLLECTION_DESCRIPTION = "Imported from Excel"
DEFAULT_TOPIC = None
MODE = "update"  # "update" or "skip"
SHEET_NAME = None  # None = active sheet

os.environ["DATABASE_URL"] = f"sqlite:///{DATABASE_PATH}"

if str(BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(BACKEND_ROOT))

from openpyxl import load_workbook
from sqlmodel import Session, select

from app.db.database import create_db_and_tables, engine
from app.models.models import Collection, Vocab


HEADER_ROW = 1

COL_GERMAN = "german"
COL_PRONUNCIATION = "pronunciation"
COL_MEANING = "meaning"
COL_EXAMPLES = "examples"


def normalize_text(value) -> str:
    if value is None:
        return ""

    return str(value).strip()


def get_first_line(value) -> str:
    text = normalize_text(value)
    lines = [line.strip() for line in text.splitlines() if line.strip()]

    if not lines:
        return ""

    return lines[0]


def clean_examples(value) -> str:
    text = normalize_text(value)

    if not text:
        return ""

    lines = text.splitlines()
    cleaned_lines = []

    for line in lines:
        stripped = line.strip()

        if not stripped:
            continue

        if stripped.lower().startswith("note"):
            break

        cleaned_lines.append(stripped)

    return "\n".join(cleaned_lines).strip()


def combine_meaning_and_pronunciation(meaning: str, pronunciation: str) -> str:
    meaning = normalize_text(meaning)
    pronunciation = normalize_text(pronunciation)

    if meaning and pronunciation:
        return f"{meaning} /{pronunciation}/"

    if meaning:
        return meaning

    if pronunciation:
        return f"/{pronunciation}/"

    return ""


def find_header_map(sheet) -> dict[str, int]:
    header_map = {}

    for cell in sheet[HEADER_ROW]:
        header = normalize_text(cell.value).lower()

        if header:
            header_map[header] = cell.column

    required_headers = [
        COL_GERMAN,
        COL_PRONUNCIATION,
        COL_MEANING,
        COL_EXAMPLES,
    ]

    missing_headers = [
        header for header in required_headers if header not in header_map
    ]

    if missing_headers:
        raise ValueError(
            "Missing required Excel headers: "
            + ", ".join(missing_headers)
            + "\nExpected headers: German, Pronunciation, meaning, Examples"
        )

    return header_map


def get_or_create_collection(
    session: Session,
    collection_name: str,
    description: Optional[str] = None,
) -> Collection:
    statement = select(Collection).where(Collection.name == collection_name)
    collection = session.exec(statement).first()

    if collection:
        return collection

    collection = Collection(
        name=collection_name,
        description=description or "Imported from Excel",
    )

    session.add(collection)
    session.commit()
    session.refresh(collection)

    return collection


def get_existing_vocab(
    session: Session,
    german: str,
    collection_id: int,
) -> Optional[Vocab]:
    statement = (
        select(Vocab)
        .where(Vocab.german == german)
        .where(Vocab.collection_id == collection_id)
    )

    return session.exec(statement).first()


def main() -> None:
    if not EXCEL_PATH.exists():
        raise FileNotFoundError(
            f"Excel file not found: {EXCEL_PATH}\n"
            f"Please put your Excel file here and rename it to: german_vocab.xlsx"
        )

    if not DATABASE_PATH.parent.exists():
        DATABASE_PATH.parent.mkdir(parents=True, exist_ok=True)

    create_db_and_tables()

    workbook = load_workbook(EXCEL_PATH, data_only=True)

    if SHEET_NAME:
        if SHEET_NAME not in workbook.sheetnames:
            raise ValueError(
                f"Sheet '{SHEET_NAME}' not found. Available sheets: {workbook.sheetnames}"
            )
        sheet = workbook[SHEET_NAME]
    else:
        sheet = workbook.active

    header_map = find_header_map(sheet)

    imported_count = 0
    updated_count = 0
    skipped_count = 0

    with Session(engine) as session:
        collection = get_or_create_collection(
            session=session,
            collection_name=COLLECTION_NAME,
            description=COLLECTION_DESCRIPTION,
        )

        if collection.id is None:
            raise RuntimeError("Collection ID is missing after creation.")

        for row_index in range(HEADER_ROW + 1, sheet.max_row + 1):
            german_raw = sheet.cell(
                row=row_index,
                column=header_map[COL_GERMAN],
            ).value

            pronunciation_raw = sheet.cell(
                row=row_index,
                column=header_map[COL_PRONUNCIATION],
            ).value

            meaning_raw = sheet.cell(
                row=row_index,
                column=header_map[COL_MEANING],
            ).value

            examples_raw = sheet.cell(
                row=row_index,
                column=header_map[COL_EXAMPLES],
            ).value

            german = get_first_line(german_raw)
            pronunciation = normalize_text(pronunciation_raw)
            meaning = normalize_text(meaning_raw)
            examples = clean_examples(examples_raw)

            if not german:
                skipped_count += 1
                continue

            vietnamese = combine_meaning_and_pronunciation(
                meaning=meaning,
                pronunciation=pronunciation,
            )

            existing_vocab = get_existing_vocab(
                session=session,
                german=german,
                collection_id=collection.id,
            )

            if existing_vocab:
                if MODE == "skip":
                    skipped_count += 1
                    continue

                existing_vocab.vietnamese = vietnamese
                existing_vocab.examples = examples
                existing_vocab.topic = DEFAULT_TOPIC
                existing_vocab.collection_id = collection.id

                session.add(existing_vocab)
                updated_count += 1
            else:
                vocab = Vocab(
                    german=german,
                    vietnamese=vietnamese,
                    examples=examples,
                    topic=DEFAULT_TOPIC,
                    collection_id=collection.id,
                    swipe_count=0,
                    mcq_correct_count=0,
                    mcq_wrong_count=0,
                )

                session.add(vocab)
                imported_count += 1

        session.commit()

    print("Excel ingestion completed.")
    print(f"Excel file: {EXCEL_PATH}")
    print(f"Database: {DATABASE_PATH}")
    print(f"Sheet: {sheet.title}")
    print(f"Collection: {COLLECTION_NAME}")
    print(f"Imported: {imported_count}")
    print(f"Updated: {updated_count}")
    print(f"Skipped: {skipped_count}")


if __name__ == "__main__":
    main()